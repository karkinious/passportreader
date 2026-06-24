import openpyxl
from openpyxl.styles import Alignment
import os
from datetime import datetime
import utils

class ExcelGenerator:
    def __init__(self, template_path='TEMPLATE.xlsx'):
        self.template_path = template_path

    def generate(self, voyage_info, crew_list):
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template file not found at {self.template_path}")

        wb = openpyxl.load_workbook(self.template_path)
        sheet = wb.active

        # Header Info (based on TEMPLATE.xlsx analysis)
        # 1. Name of ship: (4, 1) - but let's be more precise based on previous inspect
        # From inspect:
        # row 3: (None, None, None, None, 'x', 'Arrival', None, None, 'Departure', ...)
        # row 4: ('1. Name of ship :', None, None, None, '2. Port of Arrival:', ...)

        # We need to find the correct cells. Let's assume some based on common sense if not exact.
        # Actually, let's use the row/col from our inspection.

        # Arrival/Departure toggle (row 3)
        is_arrival = voyage_info.get('arrival_departure', '').lower() == 'arrival'
        if is_arrival:
            sheet.cell(row=3, column=5).value = 'x'
            sheet.cell(row=3, column=9).value = ''
        else:
            sheet.cell(row=3, column=5).value = ''
            sheet.cell(row=3, column=9).value = 'x'

        sheet.cell(row=4, column=1).value = f"1. Name of ship : {voyage_info.get('ship_name', '')}"

        port_label = "2. Port of Arrival:" if is_arrival else "2. Port of Departure:"
        sheet.cell(row=4, column=5).value = f"{port_label} {voyage_info.get('port_arrival_departure', '')}"

        date_label = "3. Date of Arrival:" if is_arrival else "3. Date of Departure:"
        sheet.cell(row=4, column=12).value = f"{date_label} {voyage_info.get('date_arrival_departure', '')}"

        sheet.cell(row=6, column=1).value = f"4. Nationality of ship: {voyage_info.get('nationality_of_ship', '')}"
        sheet.cell(row=6, column=5).value = f"5. Last Port of Call: {voyage_info.get('last_port_of_call', '')}"

        # Crew Members (Start from Row 9 based on user request)
        start_row = 9
        for i, m in enumerate(crew_list):
            row = start_row + i

            # Unmerge cells for birth and joining if they are merged in the template
            # For Birth Date, we want H-I (8-9) merged.
            # First, clear any existing merges for columns we will touch
            self._safe_unmerge(sheet, row, 8, row, 8)
            self._safe_unmerge(sheet, row, 9, row, 9)
            self._safe_unmerge(sheet, row, 10, row, 10)
            self._safe_unmerge(sheet, row, 11, row, 11)
            self._safe_unmerge(sheet, row, 12, row, 12)

            # Merge H and I for Date of Birth
            sheet.merge_cells(start_row=row, start_column=8, end_row=row, end_column=9)

            sheet.cell(row=row, column=1).value = m.get('crew_number')
            sheet.cell(row=row, column=2).value = m.get('surname')
            sheet.cell(row=row, column=3).value = m.get('given_names')
            sheet.cell(row=row, column=4).value = m.get('rank')
            sheet.cell(row=row, column=5).value = m.get('sex')
            sheet.cell(row=row, column=6).value = m.get('nationality')

            # Value for merged H-I
            sheet.cell(row=row, column=8).value = utils.format_date_display(m.get('date_of_birth'))
            sheet.cell(row=row, column=8).alignment = Alignment(horizontal='center')

            sheet.cell(row=row, column=10).value = m.get('place_of_birth')
            sheet.cell(row=row, column=11).value = utils.format_date_display(m.get('joining_date'))
            sheet.cell(row=row, column=12).value = m.get('joining_place')

            # Seaman's Book in col 13 (Nature and No.) and 14 (Expiry)
            sheet.cell(row=row, column=13).value = m.get('seamans_book_number')
            sheet.cell(row=row, column=14).value = utils.format_date_display(m.get('seamans_book_expiry'))

            # Passport in col 15 (Nature and No.) and 16 (Expiry)
            sheet.cell(row=row, column=15).value = m.get('passport_number')
            sheet.cell(row=row, column=16).value = utils.format_date_display(m.get('passport_expiry'))

        ship_name_clean = "".join(x for x in voyage_info.get('ship_name', 'Ship') if x.isalnum())
        arr_dep = voyage_info.get('arrival_departure', 'Arrival')
        output_filename = f"CrewList_{ship_name_clean}_{arr_dep}.xlsx"
        output_path = os.path.join('export', output_filename)

        wb.save(output_path)
        return output_path

    def _safe_unmerge(self, sheet, r1, c1, r2, c2):
        """Unmerges a range if it is part of merged cells."""
        # Check if the cell is part of any merged range
        # We need to unmerge the exact range that exists in merged_cells.ranges
        target_cell = sheet.cell(row=r1, column=c1).coordinate
        for merged_range in list(sheet.merged_cells.ranges):
            if target_cell in merged_range:
                sheet.unmerge_cells(str(merged_range))
                # Note: unmerging might affect other cells, but we only care about freeing our target

if __name__ == '__main__':
    print("Excel Generator ready.")
